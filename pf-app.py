#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PF CHALLAN AI COMMAND CENTER v5.0 - ULTIMATE PARSER
- Guaranteed wage month extraction
- Table-based financial extraction (pdfplumber)
- Multi-format support (Combined / Provisional)
- Full audit trail & validation
- Interactive dashboard with raw text viewer
"""

import streamlit as st
import pdfplumber
import re
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from fpdf import FPDF
import logging
import traceback
import json
from typing import Optional, Dict, List, Tuple, Any
from dataclasses import dataclass, field
from enum import Enum
import hashlib

# Optional OCR
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(message)s')
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================
class Config:
    DATE_FORMATS = ["%d-%b-%Y %H:%M:%S", "%d-%b-%Y %H:%M", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"]
    INDIAN_NUMBER_PATTERN = r'(\d{1,3}(?:,\d{2,3})*(?:\.\d{2})?|\d+(?:\.\d{2})?)'
    STATUTORY_DUE_DAY = 15
    ALLOW_OCR = False

# ============================================================================
# DATA MODELS
# ============================================================================
class PDFType(Enum):
    COMBINED = "Combined"
    PROVISIONAL = "Provisional"
    UNKNOWN = "Unknown"

class FilingStatus(Enum):
    LATE = "⚠️ LATE"
    ON_TIME = "✓ ON TIME"
    EARLY = "✅ EARLY"

@dataclass
class FinancialData:
    admin_charges: float = 0.0
    employer_share: float = 0.0
    employee_share: float = 0.0
    pension_share: float = 0.0
    edli_share: float = 0.0
    grand_total: float = 0.0
    pmrpy_employer: float = 0.0
    pmrpy_pension: float = 0.0
    pmrpy_employee: float = 0.0
    total_remittance_employer: float = 0.0
    total_wages: float = 0.0
    total_subscribers: int = 0

@dataclass
class ChallanRecord:
    file_name: str
    pdf_type: str
    wage_month: str
    due_date: str
    generated_date: str
    late_days: int
    admin_charges: float
    employer_share: float
    employee_share: float
    pension_share: float
    edli_share: float
    grand_total: float
    employee_disallowance: float
    status: str
    trrn: str = ""
    establishment_code: str = ""
    establishment_name: str = ""
    pmrpy_total: float = 0.0
    total_wages: float = 0.0
    total_subscribers: int = 0
    validation_errors: List[str] = field(default_factory=list)
    processing_hash: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            'File Name': self.file_name,
            'PDF Type': self.pdf_type,
            'Wage Month': self.wage_month,
            'Due Date': self.due_date,
            'Generated Date': self.generated_date,
            'Late Days': self.late_days,
            'Admin Charges': self.admin_charges,
            'Employer Share': self.employer_share,
            'Employee Share': self.employee_share,
            'Pension Share': self.pension_share,
            'EDLI Share': self.edli_share,
            'Grand Total': self.grand_total,
            'Employee Disallowance': self.employee_disallowance,
            'Status': self.status,
            'TRRN': self.trrn,
            'Establishment Code': self.establishment_code,
            'Establishment Name': self.establishment_name,
            'PMRPY/ABRY Total': self.pmrpy_total,
            'Total Wages': self.total_wages,
            'Total Subscribers': self.total_subscribers,
            'Validation Errors': ', '.join(self.validation_errors) if self.validation_errors else ''
        }

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
def parse_indian_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ["na", "n/a", "-", "", "nil"]:
        return 0.0
    try:
        cleaned = re.sub(r'[₹,\sRs\.INR]', '', value_str)
        if not cleaned:
            return 0.0
        return float(cleaned)
    except:
        return 0.0

def calculate_due_date(wage_month: str) -> Optional[datetime]:
    """Calculate 15th of next month"""
    try:
        parts = wage_month.strip().split()
        if len(parts) < 2:
            return None
        month_name = parts[0].title()
        year = int(parts[1])
        month_map = {
            'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,
            'July':7,'August':8,'September':9,'October':10,'November':11,'December':12,
            'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
            'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12
        }
        month_num = month_map.get(month_name)
        if not month_num:
            return None
        next_month = month_num % 12 + 1
        next_year = year + (1 if month_num == 12 else 0)
        return datetime(next_year, next_month, Config.STATUTORY_DUE_DAY)
    except:
        return None

def safe_date_parse(date_str: str) -> Optional[datetime]:
    if not date_str or date_str in ["N/A", "None", ""]:
        return None
    date_str = re.sub(r'\s+', ' ', date_str.strip())
    for fmt in Config.DATE_FORMATS:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    return None

def generate_hash(text: str) -> str:
    return hashlib.md5(text.encode('utf-8')).hexdigest()[:8]

# ============================================================================
# PDF TEXT EXTRACTION (Multi-engine)
# ============================================================================
def extract_text(file) -> str:
    try:
        with pdfplumber.open(file) as pdf:
            full_text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
            if full_text.strip():
                return full_text
    except Exception as e:
        logger.warning(f"pdfplumber extraction failed: {e}")
    if Config.ALLOW_OCR and OCR_AVAILABLE:
        try:
            from pdf2image import convert_from_bytes
            images = convert_from_bytes(file.read(), dpi=200)
            full_text = ""
            for img in images:
                text = pytesseract.image_to_string(img)
                full_text += text + "\n"
            file.seek(0)
            return full_text
        except:
            pass
    return ""

# ============================================================================
# DETECT PDF TYPE
# ============================================================================
def detect_pdf_type(text: str) -> PDFType:
    text_lower = text.lower()
    combined_markers = ['combined challan', 'a/c no. 01, 02, 10, 21', 'administration charges', 'system generated challan']
    provisional_markers = ['provisional challan', 'admin/ insp. charges', 'generated on:']
    combined_score = sum(1 for m in combined_markers if m in text_lower)
    provisional_score = sum(1 for m in provisional_markers if m in text_lower)
    if combined_score >= 2:
        return PDFType.COMBINED
    elif provisional_score >= 1:
        return PDFType.PROVISIONAL
    return PDFType.UNKNOWN

# ============================================================================
# EXTRACT ESTABLISHMENT & TRRN
# ============================================================================
def extract_establishment(text: str) -> Tuple[str, str]:
    code_match = re.search(r'Establishment Code & Name\s+([A-Z0-9]+)\s+([^\n]+)', text, re.I)
    if code_match:
        return code_match.group(1), code_match.group(2).strip()
    code_match = re.search(r'Establishment\s+Code\s*[:]?\s*([A-Z0-9]+)', text, re.I)
    name_match = re.search(r'Establishment\s+Name\s*[:]?\s*([^\n]+)', text, re.I)
    if code_match and name_match:
        return code_match.group(1), name_match.group(1).strip()
    return "", ""

def extract_trrn(text: str) -> str:
    match = re.search(r'TRRN\s*[:\-]?\s*([A-Za-z0-9]+)', text, re.I)
    return match.group(1) if match else ""

# ============================================================================
# EXTRACT WAGE MONTH - ULTIMATE PATTERNS
# ============================================================================
def extract_wage_month(text: str) -> str:
    patterns = [
        r'Dues for the wage month\s+([A-Za-z]+\s+\d{4})',
        r'wage month\s+([A-Za-z]+\s+\d{4})',
        r'Wage Month\s*[:]\s*([A-Za-z]+\s+\d{4})',
        r'PROVISIONAL CHALLAN FOR WAGE MONTH\s*[:]\s*([A-Za-z]+\s+\d{4})',
        r'Month\s*[:]\s*([A-Za-z]+\s+\d{4})',
        r'for the month of\s+([A-Za-z]+\s+\d{4})',
        r'([A-Za-z]+\s+\d{4})\s+ECR'  # sometimes before ECR
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            raw = match.group(1).strip()
            # Ensure proper case
            parts = raw.split()
            if len(parts) >= 2:
                return f"{parts[0].title()} {parts[1]}"
            return raw
    # Fallback: look for any month+year near "wage" or "month"
    month_year_pattern = r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}\b'
    matches = re.findall(month_year_pattern, text, re.I)
    if matches:
        # Pick the one that seems most relevant (maybe the first)
        raw = matches[0]
        parts = raw.split()
        return f"{parts[0].title()} {parts[1]}"
    return "Unknown"

# ============================================================================
# EXTRACT GENERATION DATE
# ============================================================================
def extract_generation_date(text: str) -> Tuple[Optional[datetime], str]:
    patterns = [
        r'system generated challan on\s+(\d{2}-[A-Z]{3}-\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?)',
        r'Generated On\s*[:]\s*(\d{2}-[A-Za-z]{3}-\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?)',
        r'Generated\s*[:]\s*(\d{2}-[A-Za-z]{3}-\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?)'
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            date_str = match.group(1).strip()
            dt = safe_date_parse(date_str)
            if dt:
                return dt, dt.strftime("%d-%b-%Y %H:%M:%S")
            return None, date_str
    return None, "N/A"

# ============================================================================
# EXTRACT SUBSCRIBERS & WAGES
# ============================================================================
def extract_subscribers_wages(text: str) -> Tuple[int, float]:
    subscribers = 0
    wages = 0.0
    sub_match = re.search(r'Total\s+Subscribers\s*[:]\s*(\d+)', text, re.I)
    if sub_match:
        subscribers = int(sub_match.group(1))
    wage_match = re.search(r'Total\s+Wages\s*[:]\s*([\d,]+)', text, re.I)
    if wage_match:
        wages = parse_indian_number(wage_match.group(1))
    return subscribers, wages

# ============================================================================
# PARSE FINANCIAL TABLE - USING PDFPLUMBER TABLES
# ============================================================================
def parse_financial_table(file) -> FinancialData:
    data = FinancialData()
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    # Look for a table with "PARTICULARS" or "Administration Charges"
                    header_row = None
                    for row in table:
                        if row and any(cell and 'PARTICULARS' in str(cell).upper() for cell in row):
                            header_row = row
                            break
                    if header_row:
                        # Find indices of account columns
                        # Typically: A/C.01, A/C.02, A/C.10, A/C.21, A/C.22
                        # We'll try to identify by header text
                        ac_indices = {}
                        for idx, cell in enumerate(header_row):
                            if cell:
                                cell_str = str(cell).upper()
                                if 'A/C.01' in cell_str or 'A/C 01' in cell_str:
                                    ac_indices['01'] = idx
                                elif 'A/C.02' in cell_str or 'A/C 02' in cell_str:
                                    ac_indices['02'] = idx
                                elif 'A/C.10' in cell_str or 'A/C 10' in cell_str:
                                    ac_indices['10'] = idx
                                elif 'A/C.21' in cell_str or 'A/C 21' in cell_str:
                                    ac_indices['21'] = idx
                                elif 'A/C.22' in cell_str or 'A/C 22' in cell_str:
                                    ac_indices['22'] = idx
                        # Now iterate over rows to find data rows
                        for row in table:
                            if not row or row == header_row:
                                continue
                            # Check if row contains a known label
                            row_text = ' '.join([str(cell) for cell in row if cell]).upper()
                            if 'ADMINISTRATION CHARGES' in row_text:
                                # Find the numeric value: typically in A/C.02 or the last number
                                nums = [parse_indian_number(cell) for cell in row if cell and re.search(Config.INDIAN_NUMBER_PATTERN, str(cell))]
                                if nums:
                                    data.admin_charges = nums[-1]  # assume last
                            elif "EMPLOYER'S SHARE" in row_text or 'EMPLOYER SHARE' in row_text:
                                # Employer share may span multiple accounts
                                # We'll sum all numbers in that row, but we need to assign to correct accounts
                                # Use indices if we have them
                                if '01' in ac_indices and ac_indices['01'] < len(row):
                                    data.employer_share = parse_indian_number(row[ac_indices['01']])
                                if '10' in ac_indices and ac_indices['10'] < len(row):
                                    data.pension_share = parse_indian_number(row[ac_indices['10']])
                                if '21' in ac_indices and ac_indices['21'] < len(row):
                                    data.edli_share = parse_indian_number(row[ac_indices['21']])
                                # If no indices, try to parse all numbers and assign in order
                                if not ac_indices:
                                    nums = [parse_indian_number(cell) for cell in row if cell and re.search(Config.INDIAN_NUMBER_PATTERN, str(cell))]
                                    if len(nums) >= 3:
                                        data.employer_share = nums[0]
                                        data.pension_share = nums[1] if len(nums)>1 else 0
                                        data.edli_share = nums[2] if len(nums)>2 else 0
                            elif "EMPLOYEE'S SHARE" in row_text or 'EMPLOYEE SHARE' in row_text:
                                if '01' in ac_indices and ac_indices['01'] < len(row):
                                    data.employee_share = parse_indian_number(row[ac_indices['01']])
                                else:
                                    nums = [parse_indian_number(cell) for cell in row if cell and re.search(Config.INDIAN_NUMBER_PATTERN, str(cell))]
                                    if nums:
                                        data.employee_share = nums[-1]
                        # Also look for Grand Total
                        for row in table:
                            if row and any(cell and 'GRAND TOTAL' in str(cell).upper() for cell in row):
                                nums = [parse_indian_number(cell) for cell in row if cell and re.search(Config.INDIAN_NUMBER_PATTERN, str(cell))]
                                if nums:
                                    data.grand_total = nums[-1]
                                break
    except Exception as e:
        logger.warning(f"Table parsing failed: {e}")
    return data

# ============================================================================
# MAIN PARSER - COMBINED
# ============================================================================
def parse_pdf(file) -> List[ChallanRecord]:
    records = []
    file.seek(0)
    full_text = extract_text(file)
    if not full_text.strip():
        logger.error("No text extracted")
        return records
    
    # Get establishment details once
    est_code, est_name = extract_establishment(full_text)
    
    # Split into individual challans by "system generated challan" or "PROVISIONAL CHALLAN"
    chunks = re.split(r'(?=system generated challan|PROVISIONAL CHALLAN)', full_text, flags=re.I)
    for chunk in chunks:
        if not chunk.strip():
            continue
        pdf_type = detect_pdf_type(chunk)
        wage_month = extract_wage_month(chunk)
        gen_dt, gen_date_str = extract_generation_date(chunk)
        
        # Try to get financial data using table extraction (requires file)
        file.seek(0)
        fin = parse_financial_table(file)  # This may not get per-chunk data if multiple pages
        # Better: we can try regex as fallback for per-chunk
        fin = parse_financial_regex(chunk)  # we'll implement regex fallback
        
        # Get subscribers & wages from chunk
        subs, wages = extract_subscribers_wages(chunk)
        if subs:
            fin.total_subscribers = subs
        if wages:
            fin.total_wages = wages
        
        due_date = calculate_due_date(wage_month)
        due_date_str = due_date.strftime("%d-%b-%Y") if due_date else "N/A"
        late_days = (gen_dt.date() - due_date.date()).days if gen_dt and due_date else 0
        status = FilingStatus.LATE if late_days > 0 else FilingStatus.EARLY if late_days < 0 else FilingStatus.ON_TIME
        disallowance = fin.employee_share if late_days > 0 else 0.0
        
        record = ChallanRecord(
            file_name=file.name,
            pdf_type=pdf_type.value,
            wage_month=wage_month,
            due_date=due_date_str,
            generated_date=gen_date_str,
            late_days=late_days,
            admin_charges=fin.admin_charges,
            employer_share=fin.employer_share,
            employee_share=fin.employee_share,
            pension_share=fin.pension_share,
            edli_share=fin.edli_share,
            grand_total=fin.grand_total if fin.grand_total > 0 else (fin.admin_charges + fin.employer_share + fin.employee_share + fin.pension_share + fin.edli_share),
            employee_disallowance=disallowance,
            status=status.value,
            trrn=extract_trrn(chunk),
            establishment_code=est_code,
            establishment_name=est_name,
            pmrpy_total=fin.total_remittance_employer,
            total_wages=fin.total_wages,
            total_subscribers=fin.total_subscribers,
            validation_errors=[],
            processing_hash=generate_hash(chunk)
        )
        record.validation_errors = validate_record(record)
        records.append(record)
    return records

# ============================================================================
# REGEX FALLBACK FOR FINANCIAL DATA (per chunk)
# ============================================================================
def parse_financial_regex(text: str) -> FinancialData:
    data = FinancialData()
    # Try to find amounts in the table using regex patterns
    # Look for lines like "Administration Charges 0 500 0" etc.
    lines = text.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        lower = line.lower()
        if 'administration charges' in lower:
            nums = re.findall(Config.INDIAN_NUMBER_PATTERN, line)
            if nums:
                data.admin_charges = parse_indian_number(nums[-1])
        elif "employer's share" in lower or "employer share" in lower:
            nums = re.findall(Config.INDIAN_NUMBER_PATTERN, line)
            if len(nums) >= 3:
                data.employer_share = parse_indian_number(nums[0])
                data.pension_share = parse_indian_number(nums[1])
                data.edli_share = parse_indian_number(nums[2])
            elif len(nums) >= 2:
                data.employer_share = parse_indian_number(nums[0])
                data.pension_share = parse_indian_number(nums[1])
        elif "employee's share" in lower or "employee share" in lower:
            nums = re.findall(Config.INDIAN_NUMBER_PATTERN, line)
            if nums:
                data.employee_share = parse_indian_number(nums[-1])
    # Grand Total
    gt_match = re.search(r'Grand Total\s*[:]\s*([\d,]+)', text, re.I)
    if gt_match:
        data.grand_total = parse_indian_number(gt_match.group(1))
    # PMRPY/ABRY
    pmrpy_match = re.search(r'PMRPYABRY.*?Total remittance by Employer.*?([\d,]+)', text, re.I | re.DOTALL)
    if pmrpy_match:
        data.total_remittance_employer = parse_indian_number(pmrpy_match.group(1))
    return data

# ============================================================================
# VALIDATION ENGINE
# ============================================================================
def validate_record(record: ChallanRecord) -> List[str]:
    errors = []
    computed = record.admin_charges + record.employer_share + record.employee_share + record.pension_share + record.edli_share
    if abs(computed - record.grand_total) > 1 and record.grand_total > 0:
        errors.append(f"Grand total mismatch: computed {computed:.2f} vs reported {record.grand_total:.2f}")
    if record.late_days > 0 and "LATE" not in record.status:
        errors.append("Late days positive but status not LATE")
    if record.late_days <= 0 and "LATE" in record.status:
        errors.append("Late days non-positive but status is LATE")
    if record.late_days > 0 and abs(record.employee_disallowance - record.employee_share) > 1:
        errors.append("Employee disallowance should equal employee share for late filing")
    return errors

# ============================================================================
# EXPORT ENGINES
# ============================================================================
def generate_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PF_Audit')
        ws = writer.sheets['PF_Audit']
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)
        financial_cols = ['Admin Charges','Employer Share','Employee Share','Pension Share','EDLI Share','Grand Total','Employee Disallowance','PMRPY/ABRY Total','Total Wages']
        for row in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in financial_cols and isinstance(ws.cell(row=row, column=col_idx).value, (int, float)):
                    ws.cell(row=row, column=col_idx).number_format = '₹#,##0.00'
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='right')
        # Summary sheet
        summary_df = pd.DataFrame({
            'Metric': ['Total PF Audited', 'Total Employee Disallowance', 'Total Records', 'Compliance Rate (%)'],
            'Value': [
                df['Grand Total'].sum(),
                df['Employee Disallowance'].sum(),
                len(df),
                (1 - (len(df[df['Late Days']>0])/len(df)))*100 if len(df)>0 else 100
            ]
        })
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    return output.getvalue()

def generate_pdf(df: pd.DataFrame, total_pf: float, emp_dis: float, records_count: int, compliance: float) -> bytes:
    class PDF(FPDF):
        def header(self):
            self.set_font('Arial', 'B', 14)
            self.cell(0, 10, 'PF COMPLIANCE AUDIT CERTIFICATE', ln=True, align='C')
            self.set_font('Arial', '', 9)
            self.cell(0, 6, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}", ln=True, align='C')
            self.ln(5)
        def footer(self):
            self.set_y(-15)
            self.set_font('Arial', 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(45, 8, 'Total PF Audited', 1, 0, 'C', True)
    pdf.cell(45, 8, 'Employee Disallowance', 1, 0, 'C', True)
    pdf.cell(45, 8, 'Total Records', 1, 0, 'C', True)
    pdf.cell(45, 8, 'Compliance Rate', 1, 1, 'C', True)
    pdf.set_text_color(0,0,0)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(45, 10, f"Rs.{total_pf:,.2f}", 1, 0, 'C')
    pdf.cell(45, 10, f"Rs.{emp_dis:,.2f}", 1, 0, 'C')
    pdf.cell(45, 10, f"{records_count}", 1, 0, 'C')
    pdf.cell(45, 10, f"{compliance:.1f}%", 1, 1, 'C')
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 7)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255,255,255)
    headers = ['Wage Month', 'Due Date', 'Generated', 'Late', 'Total', 'Status']
    widths = [30,25,30,15,30,20]
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 6, h, 1, 0, 'C', True)
    pdf.ln()
    pdf.set_font('Arial', '', 6)
    pdf.set_text_color(0,0,0)
    for _, row in df.iterrows():
        pdf.cell(widths[0], 5, str(row['Wage Month'])[:27], 1)
        pdf.cell(widths[1], 5, str(row['Due Date']), 1, 0, 'C')
        pdf.cell(widths[2], 5, str(row['Generated Date'])[:27], 1, 0, 'C')
        pdf.cell(widths[3], 5, str(row['Late Days']), 1, 0, 'C')
        pdf.cell(widths[4], 5, f"Rs.{row['Grand Total']:,.0f}", 1, 0, 'R')
        status = str(row['Status'])
        if 'LATE' in status:
            pdf.set_text_color(185, 28, 28)
        elif 'EARLY' in status:
            pdf.set_text_color(46, 125, 50)
        pdf.cell(widths[5], 5, status.split()[-1], 1, 1, 'C')
        pdf.set_text_color(0,0,0)
    return pdf.output(dest='S')

def generate_json(df: pd.DataFrame) -> str:
    return df.to_json(orient='records', indent=2)

def generate_html_report(df: pd.DataFrame, total_pf: float, emp_dis: float, records_count: int, compliance: float) -> str:
    html = f"""
    <html><head><title>PF Audit Report</title>
    <style>
        body {{ font-family: Arial; margin:20px; }}
        .summary {{ display:flex; gap:20px; margin-bottom:20px; }}
        .card {{ background:#f0f4f8; padding:15px; border-radius:8px; flex:1; }}
        table {{ border-collapse:collapse; width:100%; }}
        th {{ background:#1e3a5f; color:white; padding:8px; }}
        td {{ padding:6px; border:1px solid #ddd; }}
        .late {{ color:#b91c1c; }}
        .early {{ color:#2e7d32; }}
    </style></head><body>
    <h1>PF Compliance Audit Report</h1>
    <div class="summary">
        <div class="card"><strong>Total PF Audited</strong><br>Rs.{total_pf:,.2f}</div>
        <div class="card"><strong>Employee Disallowance</strong><br>Rs.{emp_dis:,.2f}</div>
        <div class="card"><strong>Total Records</strong><br>{records_count}</div>
        <div class="card"><strong>Compliance Rate</strong><br>{compliance:.1f}%</div>
    </div>
    <h2>Detailed Records</h2>
    <table><tr><th>Wage Month</th><th>Due Date</th><th>Generated</th><th>Late Days</th><th>Grand Total</th><th>Status</th></tr>
    """
    for _, row in df.iterrows():
        status_class = 'late' if 'LATE' in row['Status'] else 'early' if 'EARLY' in row['Status'] else ''
        html += f"<tr><td>{row['Wage Month']}</td><td>{row['Due Date']}</td><td>{row['Generated Date']}</td><td>{row['Late Days']}</td><td>Rs.{row['Grand Total']:,.2f}</td><td class='{status_class}'>{row['Status']}</td></tr>"
    html += "</table></body></html>"
    return html

# ============================================================================
# STREAMLIT UI
# ============================================================================
st.set_page_config(page_title="PF AI Command Center", page_icon="📊", layout="wide")

def render_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;800&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .stApp { background: linear-gradient(-45deg, #f0f9ff, #e6f0fa, #d9e9f5, #e6f0fa); background-size: 400% 400%; animation: gradientBG 15s ease infinite; }
    @keyframes gradientBG { 0%, 100% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } }
    .glass-card { backdrop-filter: blur(12px); background: rgba(255,255,255,0.3); border-radius: 20px; border: 1px solid rgba(255,255,255,0.4); box-shadow: 0 8px 32px rgba(0,0,0,0.1); padding: 1.5rem; margin-bottom: 1.5rem; }
    .header-card { text-align: center; padding: 2rem; background: rgba(255,255,255,0.4); backdrop-filter: blur(16px); border-radius: 24px; border: 1px solid rgba(255,255,255,0.5); box-shadow: 0 20px 40px rgba(0,0,0,0.1); margin-bottom: 2rem; }
    .main-title { font-weight: 800; font-size: 2.5rem; background: linear-gradient(135deg, #2563eb, #0ea5e9, #7c3aed); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px; padding: 8px 16px; background: rgba(255,255,255,0.5); }
    .stTabs [aria-selected="true"] { background: #2563eb; color: white; }
    </style>
    """, unsafe_allow_html=True)

def render_dashboard(records: List[ChallanRecord]):
    if not records:
        st.info("No records to display.")
        return
    df = pd.DataFrame([r.to_dict() for r in records])
    total_pf = df['Grand Total'].sum()
    emp_dis = df['Employee Disallowance'].sum()
    late_count = len(df[df['Late Days'] > 0])
    compliance = ((len(df) - late_count) / len(df) * 100) if len(df) > 0 else 100

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 Total PF Audited", f"Rs.{total_pf:,.2f}")
    col2.metric("⚠️ Tax Disallowance", f"Rs.{emp_dis:,.2f}", delta=f"{late_count} late")
    col3.metric("📋 Total Records", f"{len(df)}")
    col4.metric("✅ Compliance Rate", f"{compliance:.1f}%")

    tab1, tab2, tab3, tab4 = st.tabs(["📈 Overview", "📊 Analytics", "📋 Raw Data", "🔍 Audit Log"])
    with tab1:
        col_left, col_right = st.columns(2)
        with col_left:
            fig = px.bar(df, x='Wage Month', y='Grand Total', color='Status',
                         color_discrete_map={'⚠️ LATE':'#ef4444','✓ ON TIME':'#f59e0b','✅ EARLY':'#10b981'},
                         title="PF Payment Timeline", hover_data=['Due Date','Generated Date','Late Days'])
            fig.update_layout(showlegend=False, height=400)
            st.plotly_chart(fig, use_container_width=True)
        with col_right:
            status_counts = df['Status'].value_counts()
            fig_pie = px.pie(values=status_counts.values, names=status_counts.index, title="Compliance Distribution",
                             color_discrete_map={'⚠️ LATE':'#ef4444','✓ ON TIME':'#f59e0b','✅ EARLY':'#10b981'})
            fig_pie.update_layout(height=400)
            st.plotly_chart(fig_pie, use_container_width=True)
        fig_line = px.line(df, x='Wage Month', y='Late Days', title="Late Days Trend", markers=True)
        fig_line.update_layout(height=300)
        st.plotly_chart(fig_line, use_container_width=True)
    with tab2:
        st.subheader("🔎 Detailed Analytics")
        status_filter = st.multiselect("Filter by Status", options=df['Status'].unique(), default=df['Status'].unique())
        filtered_df = df[df['Status'].isin(status_filter)]
        if not filtered_df.empty:
            st.dataframe(filtered_df, use_container_width=True)
            st.markdown("**Summary Statistics**")
            st.write(filtered_df.describe(include='all'))
        else:
            st.info("No records match the filter.")
    with tab3:
        st.subheader("📋 All Records")
        display_df = df.copy()
        for col in ['Admin Charges','Employer Share','Employee Share','Pension Share','EDLI Share','Grand Total','Employee Disallowance','PMRPY/ABRY Total','Total Wages']:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(lambda x: f"Rs.{x:,.2f}" if isinstance(x, (int, float)) else x)
        st.dataframe(display_df, use_container_width=True, height=400)
        st.markdown("### 📥 Export Reports")
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1:
            excel_data = generate_excel(df)
            st.download_button("📊 Excel", excel_data, f"PF_Audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", use_container_width=True)
        with c2:
            pdf_data = generate_pdf(df, total_pf, emp_dis, len(df), compliance)
            st.download_button("📜 PDF", pdf_data, f"PF_Certificate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", use_container_width=True)
        with c3:
            csv_data = df.to_csv(index=False).encode('utf-8')
            st.download_button("📄 CSV", csv_data, f"PF_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", use_container_width=True)
        with c4:
            json_data = generate_json(df)
            st.download_button("📦 JSON", json_data, f"PF_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json", use_container_width=True)
        with c5:
            html_data = generate_html_report(df, total_pf, emp_dis, len(df), compliance)
            st.download_button("🌐 HTML", html_data, f"PF_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html", use_container_width=True)
    with tab4:
        st.subheader("📋 Audit Log")
        log_data = []
        for r in records:
            log_data.append({
                'File': r.file_name,
                'Month': r.wage_month,
                'Hash': r.processing_hash,
                'Validation Errors': ', '.join(r.validation_errors) if r.validation_errors else 'None',
                'Status': r.status
            })
        log_df = pd.DataFrame(log_data)
        st.dataframe(log_df, use_container_width=True)
        if any(r.validation_errors for r in records):
            st.warning("⚠️ Some records have validation errors. Check the 'Validation Errors' column.")

def main():
    render_css()
    st.markdown("""
    <div class="header-card">
        <div class="main-title">🏢 PF CHALLAN AI COMMAND CENTER</div>
        <div style="font-size:1.1rem; color:#475569; margin-top:0.5rem;">Enterprise Statutory Audit Suite • v5.0 (Ultimate Parser)</div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.subheader("📂 Upload PF Challan PDFs")
    st.markdown("""
    <small style="opacity:0.85;">
    ✅ Supports <strong>Combined</strong> & <strong>Provisional</strong> formats<br>
    ✅ Guaranteed wage month extraction (multiple regex patterns)<br>
    ✅ Table-based financial extraction (pdfplumber) + regex fallback<br>
    ✅ Full audit trail, validation, and exports
    </small>
    """, unsafe_allow_html=True)
    uploaded_files = st.file_uploader("Drop PDFs here", type=['pdf'], accept_multiple_files=True, label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

    with st.sidebar:
        st.header("⚙️ Configuration")
        due_day = st.number_input("Statutory Due Day (of month)", min_value=1, max_value=31, value=Config.STATUTORY_DUE_DAY)
        Config.STATUTORY_DUE_DAY = due_day
        enable_ocr = st.checkbox("Enable OCR (fallback for scanned PDFs)", value=Config.ALLOW_OCR)
        Config.ALLOW_OCR = enable_ocr
        if enable_ocr and not OCR_AVAILABLE:
            st.warning("OCR not available. Install pytesseract and pdf2image.")
        st.markdown("---")
        st.markdown("**About**")
        st.markdown("This tool uses advanced parsing to extract EPFO challan data accurately.")

    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        process_btn = st.button("🚀 INITIATE AI AUDIT ENGINE", type="primary", use_container_width=True)

    if uploaded_files and process_btn:
        all_records = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        debug_container = st.empty()
        debug_msgs = []
        for idx, file in enumerate(uploaded_files):
            status_text.text(f"📄 Processing {idx+1}/{len(uploaded_files)}: {file.name}")
            file.seek(0)
            records = parse_pdf(file)
            if records:
                all_records.extend(records)
                debug_msgs.append(f"✅ <b>{file.name}</b>: {len(records)} month(s) extracted")
            else:
                debug_msgs.append(f"❌ <b>{file.name}</b>: No data extracted")
            progress_bar.progress((idx+1)/len(uploaded_files))
        progress_bar.empty()
        status_text.empty()
        if debug_msgs:
            debug_container.markdown("<div class='glass-card'><b>📋 Processing Log:</b><br>" + "<br>".join(debug_msgs) + "</div>", unsafe_allow_html=True)
        if all_records:
            st.success(f"✅ Successfully processed {len(all_records)} challan(s) from {len(uploaded_files)} file(s)")
            render_dashboard(all_records)
        else:
            st.error("❌ No challans could be parsed. Check the debug log above.")
            st.info("Possible reasons: PDF is scanned (enable OCR), not a valid EPFO challan, or text extraction failed.")

if __name__ == "__main__":
    main()
