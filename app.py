import streamlit as st
import pdfplumber
import re
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import plotly.express as px
from fpdf import FPDF

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="PF AI Command Center", layout="wide", page_icon="📊")

# ---------------- ULTRA STYLISH UI (GLASSMORPHISM + ADVANCED ANIMATIONS) ----------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;800&display=swap');

/* Base Styles */
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    scroll-behavior: smooth;
}

/* Animated gradient background with floating particles */
.stApp {
    background: linear-gradient(-45deg, #f0f9ff, #e6f0fa, #d9e9f5, #e6f0fa);
    background-size: 400% 400%;
    animation: gradientBG 15s ease infinite;
    position: relative;
    overflow: hidden;
}

.stApp::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    bottom: 0;
    background: url('data:image/svg+xml;utf8,<svg width="100%" height="100%" viewBox="0 0 800 800" xmlns="http://www.w3.org/2000/svg"><circle cx="200" cy="200" r="40" fill="rgba(37,99,235,0.03)" /><circle cx="600" cy="400" r="60" fill="rgba(14,165,233,0.03)" /><circle cx="300" cy="600" r="80" fill="rgba(124,58,237,0.02)" /><circle cx="700" cy="100" r="30" fill="rgba(37,99,235,0.03)" /></svg>');
    background-size: cover;
    animation: floatParticles 30s linear infinite;
    pointer-events: none;
    z-index: 0;
}

@keyframes floatParticles {
    0% { transform: translate(0, 0) rotate(0deg); }
    50% { transform: translate(2%, 2%) rotate(1deg); }
    100% { transform: translate(0, 0) rotate(0deg); }
}

@keyframes gradientBG {
    0% { background-position: 0% 50%; }
    50% { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}

/* Dark mode background override */
@media (prefers-color-scheme: dark) {
    .stApp {
        background: linear-gradient(-45deg, #0b1a2e, #102a3c, #1a3f54, #102a3c);
        background-size: 400% 400%;
        animation: gradientBG 15s ease infinite;
    }
    .stApp::before {
        opacity: 0.2;
    }
}

/* Glassmorphism Card with glow */
.glass-card {
    backdrop-filter: blur(10px);
    background: rgba(255, 255, 255, 0.25);
    border-radius: 20px;
    border: 1px solid rgba(255, 255, 255, 0.4);
    box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
    padding: 2rem;
    margin-bottom: 2rem;
    transition: transform 0.3s ease, box-shadow 0.3s ease;
    position: relative;
    z-index: 1;
}

.glass-card:hover {
    transform: translateY(-5px);
    box-shadow: 0 12px 40px rgba(37, 99, 235, 0.2);
}

@media (prefers-color-scheme: dark) {
    .glass-card {
        background: rgba(20, 40, 60, 0.6);
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
}

/* Header Card with floating animation */
.header-card {
    text-align: center;
    padding: 2.5rem 2rem;
    background: rgba(255, 255, 255, 0.3);
    backdrop-filter: blur(12px);
    border-radius: 30px;
    border: 1px solid rgba(255, 255, 255, 0.5);
    box-shadow: 0 20px 40px rgba(0, 0, 0, 0.1);
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    animation: float 6s ease-in-out infinite;
    z-index: 1;
}

@keyframes float {
    0% { transform: translateY(0px); }
    50% { transform: translateY(-8px); }
    100% { transform: translateY(0px); }
}

.header-card::before {
    content: '';
    position: absolute;
    top: -50%;
    left: -50%;
    width: 200%;
    height: 200%;
    background: radial-gradient(circle, rgba(37, 99, 235, 0.1) 0%, transparent 70%);
    animation: rotate 20s linear infinite;
}

@keyframes rotate {
    from { transform: rotate(0deg); }
    to { transform: rotate(360deg); }
}

.main-title {
    font-weight: 800;
    font-size: 3rem;
    background: linear-gradient(135deg, #2563eb, #0ea5e9, #7c3aed);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    animation: shimmer 3s infinite;
    background-size: 200% auto;
}

@keyframes shimmer {
    0% { background-position: 0% 50%; }
    50% { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}

.subtitle {
    font-size: 1.2rem;
    font-weight: 600;
    color: #1e293b;
    opacity: 0.8;
    letter-spacing: 1px;
}

@media (prefers-color-scheme: dark) {
    .subtitle {
        color: #e2e8f0;
    }
}

/* Metric Cards with pulse on hover */
[data-testid="stMetric"] {
    background: rgba(255, 255, 255, 0.25);
    backdrop-filter: blur(8px);
    border-radius: 20px;
    padding: 1.5rem;
    border: 1px solid rgba(255, 255, 255, 0.3);
    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
    transition: all 0.3s ease;
    position: relative;
    z-index: 1;
    animation: glowPulse 4s infinite;
}

@keyframes glowPulse {
    0% { box-shadow: 0 4px 15px rgba(37, 99, 235, 0.1); }
    50% { box-shadow: 0 8px 25px rgba(37, 99, 235, 0.3); }
    100% { box-shadow: 0 4px 15px rgba(37, 99, 235, 0.1); }
}

[data-testid="stMetric"]:hover {
    transform: scale(1.03) translateY(-3px);
    background: rgba(255, 255, 255, 0.35);
    border-color: rgba(37, 99, 235, 0.5);
    animation: none;
}

[data-testid="stMetricLabel"] {
    font-weight: 600;
    font-size: 1rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    color: #1e293b;
}

[data-testid="stMetricValue"] {
    font-weight: 800;
    font-size: 2.2rem;
    color: #0f172a;
}

@media (prefers-color-scheme: dark) {
    [data-testid="stMetricLabel"] {
        color: #cbd5e1;
    }
    [data-testid="stMetricValue"] {
        color: #f1f5f9;
    }
}

/* Buttons with enhanced glow and pulse */
.stButton > button {
    background: linear-gradient(135deg, #2563eb, #0ea5e9, #7c3aed);
    background-size: 200% auto;
    color: white;
    border: none;
    border-radius: 40px;
    padding: 0.75rem 2rem;
    font-weight: 700;
    font-size: 1rem;
    letter-spacing: 0.5px;
    box-shadow: 0 4px 15px rgba(37, 99, 235, 0.3);
    transition: all 0.4s ease;
    border: 1px solid rgba(255, 255, 255, 0.2);
    backdrop-filter: blur(4px);
    width: 100%;
    position: relative;
    overflow: hidden;
}

.stButton > button::after {
    content: '';
    position: absolute;
    top: -50%;
    left: -50%;
    width: 200%;
    height: 200%;
    background: radial-gradient(circle, rgba(255,255,255,0.3) 0%, transparent 70%);
    opacity: 0;
    transition: opacity 0.4s;
}

.stButton > button:hover {
    transform: translateY(-3px) scale(1.02);
    box-shadow: 0 8px 25px rgba(37, 99, 235, 0.6);
    background-position: right center;
}

.stButton > button:hover::after {
    opacity: 0.2;
    animation: rotate 4s linear infinite;
}

.stButton > button:active {
    transform: translateY(-1px);
}

/* File uploader with scale hover */
[data-testid="stFileUploader"] {
    background: rgba(255, 255, 255, 0.2);
    backdrop-filter: blur(8px);
    border-radius: 20px;
    padding: 1.5rem;
    border: 2px dashed rgba(37, 99, 235, 0.5);
    transition: all 0.3s ease;
    z-index: 1;
    position: relative;
}

[data-testid="stFileUploader"]:hover {
    border-color: #2563eb;
    background: rgba(255, 255, 255, 0.3);
    transform: scale(1.01);
}

/* Dataframe with glass effect */
[data-testid="stDataFrame"] {
    background: rgba(255, 255, 255, 0.2);
    backdrop-filter: blur(8px);
    border-radius: 20px;
    padding: 1rem;
    border: 1px solid rgba(255, 255, 255, 0.3);
    transition: box-shadow 0.3s;
}

[data-testid="stDataFrame"]:hover {
    box-shadow: 0 8px 30px rgba(37, 99, 235, 0.15);
}

/* Plotly chart background */
.js-plotly-plot .plotly, .plotly {
    background: transparent !important;
}

/* Custom scrollbar */
::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}
::-webkit-scrollbar-track {
    background: rgba(0,0,0,0.05);
    border-radius: 10px;
}
::-webkit-scrollbar-thumb {
    background: linear-gradient(135deg, #2563eb, #0ea5e9);
    border-radius: 10px;
}
::-webkit-scrollbar-thumb:hover {
    background: #2563eb;
}

/* Footer with email link */
.footer {
    text-align: center;
    margin-top: 3rem;
    font-size: 0.9rem;
    opacity: 0.7;
    color: #1e293b;
    padding: 1rem;
    border-top: 1px solid rgba(0,0,0,0.1);
    transition: opacity 0.3s;
    z-index: 1;
    position: relative;
}

.footer:hover {
    opacity: 1;
}

.footer a {
    color: #2563eb;
    text-decoration: none;
    border-bottom: 1px dotted #2563eb;
}

.footer a:hover {
    border-bottom: 1px solid #2563eb;
}

@media (prefers-color-scheme: dark) {
    .footer {
        color: #cbd5e1;
    }
    .footer a {
        color: #60a5fa;
    }
}
</style>
""", unsafe_allow_html=True)

# ---------------- HEADER (with glass effect) ----------------
st.markdown("""
<div class="header-card">
    <div class="main-title">PF CHALLAN AI COMMAND CENTER</div>
    <div class="subtitle">Enterprise Statutory Audit Suite</div>
    <div style="margin-top:1rem; font-style:italic; opacity:0.7;">
        🌸 कर्मण्येवाधिकारस्ते मा फलेषु कदाचन 🌸
    </div>
</div>
""", unsafe_allow_html=True)

# ---------------- HELPERS (unchanged) ----------------
def safe_extract(pattern, text):
    m = re.search(pattern, text, re.I | re.M)
    if m:
        val = m.group(1).replace(",", "").strip()
        return val if val else "0"
    return "0"

def calculate_due_date(wage_month_str):
    try:
        parts = wage_month_str.split()
        month_dt = datetime.strptime(parts[0], "%B")
        year = int(parts[1])
        next_m = month_dt.month % 12 + 1
        next_y = year + (1 if month_dt.month == 12 else 0)
        return datetime(next_y, next_m, 15)
    except: return None

# ---------------- EXPORT ENGINES (unchanged) ----------------
def to_excel_pro(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PF_Audit')
        ws = writer.sheets['PF_Audit']
        h_fill, h_font = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid"), Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = h_font, h_fill, Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col])
            ws.column_dimensions[col[0].column_letter].width = max_len + 3
    return output.getvalue()

def generate_pdf_summary(df, total_pf, emp_dis):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "STATUTORY PF COMPLIANCE AUDIT CERTIFICATE", ln=True, align='C')
    pdf.set_font("Arial", '', 10)
    pdf.cell(0, 10, f"Generated For Audit Purpose | {datetime.now().strftime('%d-%m-%Y')}", ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 8, f"Total Audited: INR {total_pf:,.2f}", ln=True)
    pdf.cell(0, 8, f"Total Employee Share Disallowance: INR {emp_dis:,.2f}", ln=True)
    pdf.ln(5)

    pdf.set_font("Arial", 'B', 8); pdf.set_fill_color(30, 41, 59); pdf.set_text_color(255, 255, 255)
    w = [35, 25, 25, 15, 25, 30, 30, 30, 35, 20] 
    headers = ["Wage Month", "Due Date", "Paid Date", "Diff", "Admin", "Employer", "Employee", "Total", "Emp Disallowance", "Status"]
    for i in range(len(headers)): pdf.cell(w[i], 10, headers[i], 1, 0, 'C', True)
    pdf.ln()

    pdf.set_font("Arial", '', 7); pdf.set_text_color(0, 0, 0)
    for _, row in df.iterrows():
        pdf.cell(w[0], 8, str(row['Wage Month']), 1)
        pdf.cell(w[1], 8, str(row['Due Date']), 1, 0, 'C')
        pdf.cell(w[2], 8, str(row['Generated Date']), 1, 0, 'C')
        pdf.cell(w[3], 8, str(row['Late Days']), 1, 0, 'C')
        pdf.cell(w[4], 8, f"{row['Admin Charges']:,.2f}", 1, 0, 'R')
        pdf.cell(w[5], 8, f"{row['Employer Share']:,.2f}", 1, 0, 'R')
        pdf.cell(w[6], 8, f"{row['Employee Share']:,.2f}", 1, 0, 'R')
        pdf.cell(w[7], 8, f"{row['Grand Total']:,.2f}", 1, 0, 'R')
        pdf.cell(w[8], 8, f"{row['Employee Disallowance']:,.2f}", 1, 0, 'R')
        pdf.cell(w[9], 8, "LATE" if row['Late Days'] > 0 else "OK", 1, 1, 'C')
    
    return pdf.output(dest='S').encode('latin-1', 'replace')

# ---------------- MAIN APP ----------------
files = st.file_uploader("📂 Upload PF Challan PDFs", type="pdf", accept_multiple_files=True)

col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    run = st.button("🚀 INITIATE SYSTEM AUDIT")

if files and run:
    with st.spinner("🔮 Analyzing challans with AI precision..."):
        all_data = []
        for f in files:
            with pdfplumber.open(f) as pdf:
                text = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])
                blocks = re.split(r"(Dues for the wage month of\s*[A-Za-z]+\s*[0-9]{4})", text, flags=re.I)
                
                for i in range(1, len(blocks), 2):
                    content = blocks[i] + blocks[i+1]
                    m_match = re.search(r"wage month of\s*([A-Za-z]+)\s*([0-9]{4})", content, re.I)
                    wage_month = f"{m_match.group(1).title()} {m_match.group(2)}" if m_match else "Unknown"
                    
                    gen_date_str = safe_extract(r"system generated challan on\s*.*?(\d{2}-[A-Z]{3}-\d{4})", content).upper()
                    due_dt = calculate_due_date(wage_month)
                    
                    # Financial Data from TOTAL Column on far right ($ anchor)
                    admin = float(safe_extract(r"Administration Charges\s+.*?\s+(\d[\d,.]*)$", content))
                    employer = float(safe_extract(r"Employer'?s Share Of\s+.*?\s+(\d[\d,.]*)$", content))
                    employee = float(safe_extract(r"Employee'?s Share Of\s+.*?\s+(\d[\d,.]*)$", content))
                    total_val = float(safe_extract(r"Grand Total\s*:\s*.*?\s+(\d[\d,.]*)$", content))
                    
                    # Late Days: Negative is Early, Positive is Late
                    diff = 0
                    if due_dt and gen_date_str != "0":
                        try:
                            gen_dt = datetime.strptime(gen_date_str, "%d-%b-%Y")
                            diff = (gen_dt - due_dt).days
                        except: pass
                    
                    all_data.append({
                        "Wage Month": wage_month, "Due Date": due_dt.strftime("%d-%b-%Y") if due_dt else "N/A",
                        "Generated Date": gen_date_str, "Late Days": diff,
                        "Admin Charges": admin, "Employer Share": employer, "Employee Share": employee,
                        "Grand Total": total_val, 
                        "Employee Disallowance": employee if diff > 0 else 0.0
                    })

        if all_data:
            df = pd.DataFrame(all_data)
            st.markdown("### 📊 AUDIT DASHBOARD")
            
            # Metrics with emojis
            m1, m2, m3 = st.columns(3)
            total_pf = df['Grand Total'].sum()
            emp_dis = df['Employee Disallowance'].sum()
            m1.metric("💰 TOTAL PF PAID", f"INR {total_pf:,.2f}")
            m2.metric("⚠️ TAX DISALLOWANCE", f"INR {emp_dis:,.2f}", delta_color="inverse")
            m3.metric("⏰ LATE FILINGS", len(df[df['Late Days'] > 0]))

            st.markdown("---")
            
            # Chart
            fig = px.bar(df, x='Wage Month', y='Grand Total', color='Late Days', 
                         title="PF Payment Performance (Negative = Early)")
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#1e293b"),  # fixed color to avoid theme detection issues
                hovermode="x"
            )
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(fig, use_container_width=True)

            # Download buttons
            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "🚀 DOWNLOAD EXCEL AUDIT", 
                    to_excel_pro(df), 
                    "PF_Audit_Report.xlsx",
                    help="Download detailed audit in Excel format"
                )
            with c2:
                pdf_raw = generate_pdf_summary(df, total_pf, emp_dis)
                st.download_button(
                    "📜 DOWNLOAD PDF AUDIT TRAIL", 
                    pdf_raw, 
                    "PF_Audit_Trail.pdf", 
                    "application/pdf",
                    help="Download summary audit trail as PDF"
                )

            # Dataframe
            st.dataframe(
                df.style.format({
                    "Grand Total": "{:,.2f}", 
                    "Employee Disallowance": "{:,.2f}", 
                    "Employer Share": "{:,.2f}",
                    "Admin Charges": "{:,.2f}",
                    "Employee Share": "{:,.2f}"
                }),
                use_container_width=True,
                height=400
            )

# ---------------- FOOTER (with email and developer credit) ----------------
st.markdown("""
<div class="footer">
    © 2026 | Developed by <strong>Abhishek Jakkula</strong> | <a href="mailto:jakkulaabhishek5@gmail.com">jakkulaabhishek5@gmail.com</a><br>
    <span style="color: #2563eb;">⚡ AI-Powered Statutory Audit</span>
</div>
""", unsafe_allow_html=True)
